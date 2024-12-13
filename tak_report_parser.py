import csv
import time
import tkinter
import customtkinter as ctk
from tkinter import filedialog, messagebox
import os
import requests
import xml.etree.ElementTree as ET
import zipfile
from io import BytesIO
from cryptography.hazmat.primitives.serialization import pkcs12
from cryptography.hazmat.primitives.serialization import Encoding, PrivateFormat, NoEncryption
from cryptography.hazmat.backends import default_backend
from openpyxl import Workbook
from datetime import datetime, timezone, timedelta
import re
import shutil
import threading
from pygeodesy import ellipsoidalVincenty as ev
import Home_Page
import urllib3

# Suppress InsecureRequestWarning messages for faster parsing
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Set up customtkinter
ctk.set_appearance_mode("System")
ctk.set_default_color_theme("dark-blue")

class TAKReportGUI(ctk.CTk):
    def __init__(self):
        super().__init__()

        # Immediately define a stable repository folder and file
        # so that previous connections can be read anytime.
        app_dir = os.path.dirname(os.path.abspath(__file__))
        self.repository_folder = os.path.join(app_dir, 'ServerConnections')
        self.repository_file = os.path.join(self.repository_folder, 'connections.csv')
        os.makedirs(self.repository_folder, exist_ok=True)

        # User input variables
        self.pfx_file_path = ctk.StringVar()
        self.pfx_password = ctk.StringVar()
        self.base_url = ctk.StringVar()
        self.port_number = ctk.StringVar()
        self.template_path = ctk.StringVar()
        self.output_folder = ctk.StringVar()  # Chosen by the user
        self.timezone_selection = ctk.StringVar(value="EST")
        self.start_datetime_str = ctk.StringVar()
        self.output_option = ctk.StringVar(value="Combined Workbook")

        # We'll define directories after the user chooses the output folder
        self.output_parent_folder = None
        self.media_folder = None
        self.combined_reports_path = None

        # Setup frames
        self.setup_frames()
        self.mainloop()

    def browse_pfx_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("PFX/P12 Files", "*.pfx *.p12")])
        if file_path:
            self.pfx_file_path.set(file_path)

    def browse_template_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("XML Files", "*.xml")])
        if file_path:
            self.template_path.set(file_path)

    def browse_output_folder(self):
        folder_selected = filedialog.askdirectory()
        if folder_selected:
            self.output_folder.set(folder_selected)
            # As soon as the user selects an output folder, ensure directories are created.
            self.ensure_output_directories(folder_selected)

    def ensure_output_directories(self, output_folder):
        self.output_parent_folder = output_folder
        self.media_folder = os.path.join(self.output_parent_folder, 'TAK Reports Media')
        self.combined_reports_path = os.path.join(self.output_parent_folder, 'combined_reports.xml')

        # Create directories
        os.makedirs(self.output_parent_folder, exist_ok=True)
        os.makedirs(self.media_folder, exist_ok=True)

    def setup_frames(self):
        input_frame = ctk.CTkFrame(self)
        input_frame.grid(row=0, column=0, padx=20, pady=20, sticky="nsew")
        input_frame.grid_columnconfigure(0, weight=1)
        input_frame.grid_columnconfigure(1, weight=1)

        title_label = ctk.CTkLabel(input_frame, text="TAK Report Parser", font=("Arial", 36, "bold"))
        title_label.grid(row=0, column=0, columnspan=3, pady=(10, 20))

        ctk.CTkLabel(input_frame, text="TAK Certificate File Path:").grid(row=1, column=0, padx=10, pady=10, sticky="e")
        ctk.CTkEntry(input_frame, textvariable=self.pfx_file_path, width=300).grid(row=1, column=1, padx=10, pady=10)
        ctk.CTkButton(input_frame, text="Browse", command=self.browse_pfx_file).grid(row=1, column=2, padx=10, pady=10)

        ctk.CTkLabel(input_frame, text="TAK Certificate Password:").grid(row=2, column=0, padx=10, pady=10, sticky="e")
        ctk.CTkEntry(input_frame, textvariable=self.pfx_password, show="*", width=300).grid(row=2, column=1, padx=10, pady=10)

        ctk.CTkLabel(input_frame, text="Base URL:").grid(row=3, column=0, padx=10, pady=10, sticky="e")
        ctk.CTkEntry(input_frame, textvariable=self.base_url, width=300).grid(row=3, column=1, padx=10, pady=10)

        ctk.CTkLabel(input_frame, text="Port Number:").grid(row=4, column=0, padx=10, pady=10, sticky="e")
        ctk.CTkEntry(input_frame, textvariable=self.port_number, width=100).grid(row=4, column=1, padx=10, pady=10)

        ctk.CTkLabel(input_frame, text="TAK Report XML Template Path:").grid(row=5, column=0, padx=10, pady=10, sticky="e")
        ctk.CTkEntry(input_frame, textvariable=self.template_path, width=300).grid(row=5, column=1, padx=10, pady=10)
        ctk.CTkButton(input_frame, text="Browse", command=self.browse_template_file).grid(row=5, column=2, padx=10, pady=10)

        ctk.CTkLabel(input_frame, text="Output Folder:").grid(row=6, column=0, padx=10, pady=10, sticky="e")
        ctk.CTkEntry(input_frame, textvariable=self.output_folder, width=300).grid(row=6, column=1, padx=10, pady=10)
        ctk.CTkButton(input_frame, text="Browse", command=self.browse_output_folder).grid(row=6, column=2, padx=10, pady=10)

        ctk.CTkLabel(input_frame, text="Timezone:").grid(row=7, column=0, padx=10, pady=10, sticky="e")
        ctk.CTkOptionMenu(input_frame, values=["EST", "CST", "MST", "PST"], command=self.set_timezone, variable=self.timezone_selection).grid(row=7, column=1, padx=10, pady=10, sticky="w")

        ctk.CTkLabel(input_frame, text="Start Date/Time (YYYY-MM-DD HH:MM:SS):").grid(row=8, column=0, padx=10, pady=10, sticky="e")
        ctk.CTkEntry(input_frame, textvariable=self.start_datetime_str, width=300).grid(row=8, column=1, padx=10, pady=10)

        ctk.CTkLabel(input_frame, text="CSV Output Option:").grid(row=9, column=0, padx=10, pady=10, sticky="e")
        ctk.CTkOptionMenu(input_frame, values=["Combined Workbook", "Separate Workbooks"], variable=self.output_option).grid(row=9, column=1, padx=10, pady=10, sticky="w")

        ctk.CTkButton(input_frame, text="Use Previous Connection", command=self.select_previous_connection, width=120).grid(row=10, column=0, columnspan=3, padx=20, pady=20, sticky="n")

        ctk.CTkButton(input_frame, text="Start Parsing", command=self.start_parsing, width=120).grid(row=11, column=0, columnspan=3, padx=20, pady=20, sticky="n")

        return_button = ctk.CTkButton(self, text="Return to Home Page", command=self.return_to_home)
        return_button.grid(pady=20)

    def set_timezone(self, selection):
        self.timezone_selection.set(selection)

    def start_parsing(self):
        pfx_file = self.pfx_file_path.get()
        password = self.pfx_password.get()
        base_url = self.base_url.get()
        port = self.port_number.get()
        template_path = self.template_path.get()
        output_folder = self.output_folder.get()
        timezone = self.timezone_selection.get()
        start_datetime_str = self.start_datetime_str.get()
        output_option = self.output_option.get()

        if not all([pfx_file, password, base_url, port, template_path, start_datetime_str, output_folder]):
            messagebox.showerror("Input Error", "All fields including output folder must be completed.")
            return

        try:
            self.start_datetime = datetime.strptime(start_datetime_str, '%Y-%m-%d %H:%M:%S')
        except ValueError:
            messagebox.showerror("Input Error", "Invalid Start Date/Time format. Please use YYYY-MM-DD HH:MM:SS")
            return

        confirmation = messagebox.askyesno(
            "Confirmation",
            f"Proceed with the following details?\n\n"
            f"PFX File: {pfx_file}\n"
            f"Base URL: {base_url}\n"
            f"Port: {port}\n"
            f"Template Path: {template_path}\n"
            f"Output Folder: {output_folder}\n"
            f"Timezone: {timezone}\n"
            f"Start Date/Time: {start_datetime_str}\n"
            f"Output Option: {output_option}"
        )
        if confirmation:
            # Directories are ensured when output folder was chosen. Just run the process.
            self.run_process()

    def read_connections(self):
        if os.path.exists(self.repository_file):
            with open(self.repository_file, mode='r', newline='') as file:
                reader = csv.DictReader(file)
                return list(reader)
        return []

    def write_connection(self, ticket_number, base_url, port, pfx_file, password, template_path, output_folder):
        fieldnames = ['Ticket Number', 'Base URL', 'Port', 'PFX File', 'Password', 'Template Path', 'Output Folder']
        file_exists = os.path.exists(self.repository_file)

        with open(self.repository_file, mode='a', newline='') as file:
            writer = csv.DictWriter(file, fieldnames=fieldnames)
            if not file_exists:
                writer.writeheader()
            writer.writerow({
                'Ticket Number': ticket_number,
                'Base URL': base_url,
                'Port': port,
                'PFX File': pfx_file,
                'Password': password,
                'Template Path': template_path,
                'Output Folder': output_folder
            })

    def get_next_ticket_number(self):
        connections = self.read_connections()
        if connections:
            ticket_numbers = [int(conn['Ticket Number']) for conn in connections]
            return max(ticket_numbers) + 1
        return 1

    def select_previous_connection(self):
        connections = self.read_connections()
        if not connections:
            messagebox.showinfo("No Connections", "No previous connections found.")
            return

        ticket_numbers = [f"Ticket {conn['Ticket Number']}: {conn['Base URL']}" for conn in connections]
        selected_ticket = self.prompt_ticket_selection(ticket_numbers)

        selected_connection = next((conn for conn in connections if f"Ticket {conn['Ticket Number']}" in selected_ticket), None)

        if selected_connection:
            self.pfx_file_path.set(selected_connection['PFX File'])
            self.pfx_password.set(selected_connection['Password'])
            self.base_url.set(selected_connection['Base URL'])
            self.port_number.set(selected_connection['Port'])
            self.template_path.set(selected_connection['Template Path'])
            self.output_folder.set(selected_connection.get('Output Folder', ''))
            # Now that we have output folder, ensure directories
            if self.output_folder.get():
                self.ensure_output_directories(self.output_folder.get())

    def prompt_ticket_selection(self, ticket_numbers):
        selection_window = ctk.CTkToplevel(self)
        selection_window.title("Select Ticket")

        selection_window.geometry("500x500")
        selection_window.grid_columnconfigure(0, weight=1)

        selection_window.lift()
        selection_window.grab_set()

        ctk.CTkLabel(selection_window, text="Select a ticket number:").grid(row=0, column=0, pady=10)

        listbox_frame = ctk.CTkFrame(selection_window)
        listbox_frame.grid(row=1, column=0, padx=10, pady=10)
        lb = tkinter.Listbox(listbox_frame, height=12, width=50, font=("Arial", 20))
        lb.pack()

        for ticket in ticket_numbers:
            lb.insert("end", ticket)

        def on_select():
            try:
                selected_ticket = lb.get(lb.curselection())
                selection_window.destroy()
                self.selected_ticket = selected_ticket
            except tkinter.TclError:
                messagebox.showerror("Selection Error", "Please select a ticket.")
                return

        select_button = ctk.CTkButton(selection_window, text="Select", command=on_select)
        select_button.grid(row=2, column=0, pady=10)

        self.wait_window(selection_window)
        return self.selected_ticket

    def validate_file_path(self, path):
        return os.path.isfile(path)

    def validate_password(self, pfx_file, password):
        try:
            with open(pfx_file, 'rb') as f:
                pfx_data = f.read()
            pkcs12.load_key_and_certificates(pfx_data, password.encode(), backend=default_backend())
            return True
        except Exception:
            return False

    def validate_base_url(self, url):
        return '.' in url

    def validate_port(self, port):
        return port.isdigit() and len(port) == 4

    def pfx_to_pem(self, pfx_file, password):
        try:
            with open(pfx_file, 'rb') as f:
                pfx_data = f.read()
            private_key, certificate, additional_certificates = pkcs12.load_key_and_certificates(
                pfx_data, password.encode(), backend=default_backend()
            )

            cert_pem = certificate.public_bytes(Encoding.PEM).decode()
            key_pem = private_key.private_bytes(
                Encoding.PEM,
                PrivateFormat.PKCS8,
                NoEncryption()
            ).decode()

            cert_file = pfx_file.replace('.p12', '_cert.pem').replace('.pfx', '_cert.pem')
            key_file = pfx_file.replace('.p12', '_key.pem').replace('.pfx', '_key.pem')

            with open(cert_file, 'w') as f:
                f.write(cert_pem)

            with open(key_file, 'w') as f:
                f.write(key_pem)

            return cert_file, key_file
        except Exception as e:
            messagebox.showerror("Conversion Error", f"Failed to convert PFX to PEM: {str(e)}")
            return None, None

    def new_server_connection(self):
        pfx_file = self.pfx_file_path.get()
        password = self.pfx_password.get()
        base_url = self.base_url.get()
        port = self.port_number.get()
        template_path = self.template_path.get()
        output_folder = self.output_folder.get()

        if not self.validate_file_path(pfx_file):
            messagebox.showerror("Error", "Invalid PFX file path. Please check the path and try again.")
            return None, None, None

        if not self.validate_password(pfx_file, password):
            messagebox.showerror("Error", "Invalid password for the PFX file. Please try again.")
            return None, None, None

        if not self.validate_base_url(base_url):
            messagebox.showerror("Error", "Invalid base URL. Please enter a valid URL.")
            return None, None, None

        if not self.validate_port(port):
            messagebox.showerror("Error", "Invalid port number. Please enter a valid 4-digit port number.")
            return None, None, None

        if not output_folder:
            messagebox.showerror("Error", "No output folder selected. Please choose an output folder.")
            return None, None, None

        metadata_url = f'https://{base_url}:{port}/Marti/api/files/metadata?'
        file_url_template = f'https://{base_url}:{port}/Marti/api/files/{{hash}}'
        ticket_number = self.get_next_ticket_number()

        try:
            cert_file, key_file = self.pfx_to_pem(pfx_file, password)
            if cert_file and key_file:
                print("PFX to PEM conversion successful")
            else:
                raise ValueError("PEM conversion failed")
        except ValueError as e:
            messagebox.showerror("Conversion Error", f"PEM Conversion failed: {e}")
            return None, None, None

        ssl_cert = (cert_file, key_file)
        self.write_connection(ticket_number, base_url, port, pfx_file, password, template_path, output_folder)

        return metadata_url, file_url_template, ssl_cert

    def run_process(self):
        if hasattr(self, 'selected_ticket'):
            connections = self.read_connections()
            selected_connection = None
            for conn in connections:
                if f"Ticket {conn['Ticket Number']}" in self.selected_ticket:
                    selected_connection = conn
                    break

            if selected_connection:
                base_url = selected_connection['Base URL']
                port = selected_connection['Port']
                pfx_file = selected_connection['PFX File']
                password = selected_connection['Password']
                output_folder = selected_connection.get('Output Folder', '')

                self.pfx_password.set(password)

                if not self.validate_password(pfx_file, password):
                    messagebox.showerror("Invalid Password", "Invalid password. Please try again.")
                    return

                try:
                    cert_file, key_file = self.pfx_to_pem(pfx_file, password)
                    messagebox.showinfo("Conversion Successful", "PFX to PEM conversion successful.")
                except ValueError as e:
                    messagebox.showerror("Conversion Failed", f"Conversion failed: {e}")
                    return

                ssl_cert = (cert_file, key_file)
                metadata_url = f'https://{base_url}:{port}/Marti/api/files/metadata'
                file_url_template = f'https://{base_url}:{port}/Marti/api/files/{{hash}}'

                if not output_folder:
                    messagebox.showerror("No Output Folder", "No output folder was stored in this connection. Please select a folder and create a new connection.")
                    return
                # Directories ensured when output folder was chosen (or after selecting previous connection)
            else:
                messagebox.showerror("Invalid Selection", "Ticket number not found. You will need to connect to a new server.")
                return
        else:
            metadata_url, file_url_template, ssl_cert = self.new_server_connection()
            if metadata_url is None:
                return
            # Directories are ensured after output folder was chosen

        template_path = self.template_path.get()
        timezone = self.timezone_selection.get()

        tz_offsets = {'EST': -5, 'CST': -6, 'MST': -7, 'PST': -8}
        tz_offset = tz_offsets.get(timezone, -5)
        tz_name = timezone

        self.templates = self.parse_template(template_path)
        self.fetch_reports(metadata_url, file_url_template, ssl_cert)

    def fetch_reports(self, metadata_url, file_url_template, ssl_cert):
        reports = []
        combined_reports = ET.Element('CombinedReports')
        processed_hashes = set()

        try:
            response = requests.get(metadata_url, cert=ssl_cert, verify=False, timeout=60)
            response.raise_for_status()
            metadata = response.json()
            messagebox.showinfo("Metadata Received", f"Metadata received with {len(metadata['data'])} entries.")
        except requests.exceptions.RequestException as e:
            messagebox.showerror("Request Failed", f"Request failed: {e}")
            return

        if 'data' not in metadata or not metadata['data']:
            messagebox.showinfo("No Reports", "No reports available on the server.")
            return

        progress_window = ctk.CTkToplevel(self)
        progress_window.title("Processing Reports")
        progress_window.geometry("500x150")
        progress_window.grab_set()

        progress_label = ctk.CTkLabel(progress_window, text="Starting...")
        progress_label.pack(pady=10)

        progress_bar = ctk.CTkProgressBar(progress_window, orientation='horizontal', width=300)
        progress_bar.pack(pady=20)
        progress_bar.set(0)

        self.total_entries = len(metadata['data'])
        self.current_index = 0
        self.metadata = metadata
        self.file_url_template = file_url_template
        self.ssl_cert = ssl_cert
        self.reports = reports
        self.combined_reports = combined_reports
        self.processed_hashes = processed_hashes
        self.folder_counter = 1
        self.progress_bar = progress_bar
        self.progress_label = progress_label
        self.progress_window = progress_window

        self.process_next_entry()

    def process_next_entry(self):
        if self.current_index >= self.total_entries:
            if len(self.combined_reports) > 0:
                combined_tree = ET.ElementTree(self.combined_reports)
                combined_tree.write(self.combined_reports_path, encoding='utf-8', xml_declaration=True)
                messagebox.showinfo("Reports Saved", f"Combined XML reports saved to {self.combined_reports_path}.")
            else:
                messagebox.showinfo("No Reports", "No citrap reports found to save.")

            self.progress_window.destroy()

            try:
                timezone = self.timezone_selection.get()
                tz_offsets = {'EST': -5, 'CST': -6, 'MST': -7, 'PST': -8}
                tz_offset = tz_offsets.get(timezone, -5)
                tz_name = timezone
                self.parse_reports(self.templates, self.reports, tz_offset, tz_name)
            except Exception as e:
                messagebox.showerror("Error", f"An error occurred: {e}")
            return

        file_info = self.metadata['data'][self.current_index]
        index = self.current_index + 1

        print(f"Processing entry {index}/{self.total_entries}: {file_info['Name']}")

        if 'citrap' not in file_info.get('Keywords', '').lower():
            print(f"Skipping non-citrap report: {file_info['Name']}")
            progress = index / self.total_entries
            self.progress_bar.set(progress)
            self.progress_label.configure(text=f"Processing {index}/{self.total_entries}: Skipped")
            self.current_index += 1
            self.after(0, self.process_next_entry)
            return

        report_hash = file_info['Hash']
        if report_hash in self.processed_hashes:
            print(f"Skipping already processed report: {file_info['Name']}")
            progress = index / self.total_entries
            self.progress_bar.set(progress)
            self.progress_label.configure(text=f"Processing {index}/{self.total_entries}: Skipped")
            self.current_index += 1
            self.after(0, self.process_next_entry)
            return

        self.processed_hashes.add(report_hash)

        retries = 3
        while retries > 0:
            try:
                file_url = self.file_url_template.format(hash=report_hash)
                file_response = requests.get(file_url, cert=self.ssl_cert, verify=False, timeout=60)
                file_response.raise_for_status()

                if file_info['MimeType'] == 'application/zip':
                    try:
                        with zipfile.ZipFile(BytesIO(file_response.content)) as z:
                            for file_name in z.namelist():
                                if file_name.endswith('.xml'):
                                    with z.open(file_name) as xml_file:
                                        try:
                                            report_tree = ET.ElementTree(ET.parse(xml_file).getroot())
                                            self.combined_reports.append(report_tree.getroot())
                                            self.reports.append(report_tree.getroot())
                                        except ET.ParseError as e:
                                            messagebox.showerror("XML Parsing Error", f"Failed to parse XML from {file_name}: {e}")
                                else:
                                    folder_name = f"{self.folder_counter:02d}"
                                    media_folder_path = os.path.join(self.media_folder, folder_name)
                                    os.makedirs(os.path.dirname(os.path.join(media_folder_path, file_name)), exist_ok=True)
                                    media_path = os.path.join(media_folder_path, file_name)
                                    with z.open(file_name) as media_file:
                                        with open(media_path, 'wb') as out_file:
                                            shutil.copyfileobj(media_file, out_file)
                                    print(f"Saved media file: {file_name} to {media_path}")
                            self.folder_counter += 1
                    except zipfile.BadZipFile:
                        break
                elif file_info['MimeType'] == 'application/xml' or file_info['Name'].endswith('.xml'):
                    try:
                        report_tree = ET.ElementTree(ET.fromstring(file_response.content))
                        self.combined_reports.append(report_tree.getroot())
                        self.reports.append(report_tree.getroot())
                    except ET.ParseError as e:
                        messagebox.showerror("XML Parsing Error", f"Failed to parse XML: {e}")
                else:
                    folder_name = f"{self.folder_counter:02d}"
                    media_folder_path = os.path.join(self.media_folder, folder_name)
                    os.makedirs(os.path.dirname(os.path.join(media_folder_path, file_info['Name'])), exist_ok=True)
                    media_path = os.path.join(media_folder_path, file_info['Name'])
                    with open(media_path, 'wb') as out_file:
                        out_file.write(file_response.content)
                    print(f"Saved media file: {file_info['Name']} to {media_path}")
                    self.folder_counter += 1

                break
            except requests.exceptions.RequestException as e:
                retries -= 1
                messagebox.showerror("Fetching Error", f"Failed to fetch report from {file_url}: {e}")
                if retries > 0:
                    print("Retrying...")
                    time.sleep(5)
                else:
                    messagebox.showerror("Retry Failed", f"Giving up on {file_info['Name']} after 3 attempts.")

        progress = index / self.total_entries
        self.progress_bar.set(progress)
        self.progress_label.configure(text=f"Processing {index}/{self.total_entries}: {file_info['Name']}")

        self.current_index += 1
        self.after(0, self.process_next_entry)

    def parse_template(self, template_path):
        try:
            with open(template_path, 'r', encoding='utf-8') as file:
                content = file.read().lstrip()
                tree = ET.ElementTree(ET.fromstring(content))
        except ET.ParseError as e:
            messagebox.showerror("Template Error", f"Failed to parse the XML template: {e}")
            return {}
        root = tree.getroot()
        templates = {}

        for report in root.findall('report'):
            report_type = report.get('type')
            fields = [
                {'xml_path': '.', 'csv_header': 'Report Type', 'attribute': 'type', 'type': 'attribute'},
                {'xml_path': '.', 'csv_header': 'Reporter Callsign', 'attribute': 'userCallsign', 'type': 'attribute'},
                {'xml_path': '.', 'csv_header': 'Date/Time', 'attribute': 'dateTime', 'type': 'attribute'},
                {'xml_path': '.', 'csv_header': 'Location', 'attribute': 'location', 'type': 'location'},
                {'xml_path': '.', 'csv_header': 'Report Title', 'attribute': 'title', 'type': 'attribute'}
            ]

            for section in report.findall('.//section'):
                fields.append({
                    'xml_path': '.', 'csv_header': section.get('title'), 'attribute': None, 'type': 'section'
                })
                for option in section.findall('option'):
                    opt_type = option.get('type')
                    title = option.get('title')
                    if opt_type == 'string':
                        fields.append({'xml_path': title, 'csv_header': title, 'attribute': 'value', 'type': 'text'})
                    elif opt_type == 'checkbox':
                        fields.append({'xml_path': title, 'csv_header': title, 'attribute': 'value', 'type': 'checkbox'})
                    elif opt_type == 'dateTime':
                        fields.append({'xml_path': title, 'csv_header': title, 'attribute': 'value', 'type': 'date'})
                    elif opt_type == 'geometry':
                        fields.append({'xml_path': title, 'csv_header': title, 'attribute': 'value', 'type': 'geometry'})
                    elif opt_type == 'number' and option.get('unitOptions') is not None:
                        fields.append({'xml_path': title, 'csv_header': title, 'attribute': 'value', 'type': 'number_with_units', 'unit': option.get('unitValue')})
                    elif opt_type == 'number':
                        fields.append({'xml_path': title, 'csv_header': title, 'attribute': 'value', 'type': 'number'})
                    elif opt_type == 'rangeBearing':
                        fields.append({'xml_path': title, 'csv_header': title, 'attribute': 'value', 'type': 'range'})
                    elif opt_type == 'route':
                        fields.append({'xml_path': title, 'csv_header': title, 'attribute': 'value', 'type': 'route'})
                    elif opt_type == 'time':
                        fields.append({'xml_path': title, 'csv_header': title, 'attribute': 'value', 'type': 'time'})

            for list_ in report.findall('.//list'):
                list_title = list_.get('title')
                if list_.get('multiple') == 'true':
                    fields.append({'xml_path': list_title, 'csv_header': list_title, 'attribute': None, 'type': 'multi-select'})
                else:
                    fields.append({'xml_path': list_title, 'csv_header': list_title, 'attribute': None, 'type': 'list'})

            templates[report_type] = fields

        return templates

    def process_reports_for_type(self, reports, report_type, fields, main_sheet, duplicate_sheet, tz_offset, tz_name):
        def safe_value(value):
            if value is None or (isinstance(value, str) and value.strip() == ''):
                return ' '
            else:
                return value

        seen_datetimes = {}
        duplicates = []

        for report in reports:
            if report.get('type') == report_type:
                row = []
                report_time = None

                for field in fields:
                    if field['csv_header'] == 'Date/Time' and field['attribute'] == 'dateTime':
                        zulu_time = report.get(field['attribute'])
                        local_time = self.convert_zulu_to_timezone(zulu_time, tz_offset, tz_name)
                        row.append(safe_value(local_time))
                        try:
                            report_time = datetime.strptime(local_time, '%Y-%m-%d %H:%M:%S')
                        except ValueError:
                            continue
                    elif field['type'] == 'location':
                        location_str = report.get(field['attribute'])
                        lat, lon = self.extract_latlong_from_location(location_str)
                        if lat is not None and lon is not None:
                            mgrs_coord = self.convert_latlong_to_mgrs(lat, lon)
                            row.append(safe_value(mgrs_coord))
                        else:
                            row.append(' ')
                    elif field['type'] == 'geometry':
                        found_geometry = False
                        for section in report.findall('.//section'):
                            for option in section.findall('option'):
                                if option.get('title') == field['csv_header'] and option.get('type') == 'geometry':
                                    location_str = option.get('value')
                                    lat, lon = self.extract_latlong_from_location(location_str)
                                    if lat is not None and lon is not None:
                                        mgrs_coord = self.convert_latlong_to_mgrs(lat, lon)
                                        row.append(safe_value(mgrs_coord))
                                    else:
                                        row.append(' ')
                                    found_geometry = True
                                    break
                            if found_geometry:
                                break
                        if not found_geometry:
                            row.append(' ')
                    elif field['type'] == 'section':
                        row.append(' ')
                    elif field['type'] == 'text':
                        element = report.find('.//option[@title="' + field['xml_path'] + '"]')
                        value = element.get(field['attribute']) if element is not None else None
                        row.append(safe_value(value))
                    elif field['type'] == 'checkbox':
                        element = report.find('.//option[@title="' + field['xml_path'] + '"]')
                        value = 'X' if element is not None and element.get('value') == 'True' else ' '
                        row.append(value)
                    elif field['type'] == 'date':
                        element = report.find('.//option[@title="' + field['xml_path'] + '"]')
                        zulu_time = element.get('value') if element is not None else None
                        if zulu_time:
                            local_time = self.convert_zulu_to_timezone(zulu_time, tz_offset, tz_name)
                            row.append(safe_value(local_time))
                        else:
                            row.append(' ')
                    elif field['type'] == 'number_with_units':
                        element = report.find('.//option[@title="' + field['xml_path'] + '"]')
                        if element is not None:
                            value = element.get('value')
                            unit = element.get('unitValue')
                            combined_value = f"{value} {unit}" if value and unit else None
                            row.append(safe_value(combined_value))
                        else:
                            row.append(' ')
                    elif field['type'] == 'number':
                        element = report.find('.//option[@title="' + field['xml_path'] + '"]')
                        value = element.get('value') if element is not None else None
                        row.append(safe_value(value))
                    elif field['type'] == 'range':
                        element = report.find('.//option[@title="' + field['xml_path'] + '"]')
                        value = element.get('value') if element is not None else None
                        row.append(safe_value(value))
                    elif field['type'] == 'route':
                        element = report.find('.//option[@title="' + field['xml_path'] + '"]')
                        value = element.get('value') if element is not None else None
                        row.append(safe_value(value))
                    elif field['type'] == 'time':
                        element = report.find('.//option[@title="' + field['xml_path'] + '"]')
                        zulu_time = element.get('value') if element is not None else None
                        if zulu_time:
                            local_time = self.convert_zulu_to_timezone(zulu_time, tz_offset, tz_name)
                            row.append(safe_value(local_time))
                        else:
                            row.append(' ')
                    elif field['type'] == 'list':
                        selected_option = None
                        list_element = report.find('.//list[@title="' + field['csv_header'] + '"]')
                        if list_element is not None:
                            for option_el in list_element.findall('.//option'):
                                if option_el.get('selected') == 'true':
                                    selected_option = option_el.get('title')
                                    break
                        row.append(safe_value(selected_option))
                    elif field['type'] == 'multi-select':
                        selected_options = []
                        list_element = report.find('.//list[@title="' + field['csv_header'] + '"]')
                        if list_element is not None:
                            for option_el in list_element.findall('.//option'):
                                if option_el.get('selected') == 'true':
                                    selected_options.append(option_el.get('title'))
                        multi_value = ', '.join(selected_options) if selected_options else None
                        row.append(safe_value(multi_value))
                    elif field['attribute']:
                        value = report.get(field['attribute'])
                        row.append(safe_value(value))
                    else:
                        element = report.find(field['xml_path'])
                        value = element.text if element is not None else None
                        row.append(safe_value(value))

                if report_time and report_time < self.start_datetime:
                    continue
                if report_time:
                    if report_time in seen_datetimes:
                        duplicates.append((report_time, row))
                    else:
                        seen_datetimes[report_time] = row

        sorted_unique_rows = sorted(seen_datetimes.items())
        sorted_duplicates = sorted(duplicates, key=lambda x: x[0])

        for _, row in sorted_unique_rows:
            main_sheet.append(row)

        for _, row in sorted_duplicates:
            duplicate_sheet.append(row)

    def delete_columns(self, workbook):
        # Only keep unconditional deletion of the 6th column
        for sheet in workbook.worksheets:
            try:
                sheet.delete_cols(6)
            except Exception as e:
                print(f"Failed to delete column F in sheet '{sheet.title}': {e}")

    def convert_latlong_to_mgrs(self, lat_str, lon_str):
        try:
            lat = float(lat_str)
            lon = float(lon_str)
            latlon = ev.LatLon(lat, lon)
            mgrs_coord = latlon.toMgrs()
            return str(mgrs_coord)
        except Exception as e:
            print(f"Failed to convert Lat/Long to MGRS: {e}")
            return f"{lat_str}, {lon_str}"

    def extract_latlong_from_location(self, location_str):
        if location_str is None:
            print("No location data found; skipping extraction.")
            return None, None

        try:
            match = re.search(r"POINT \(([-\d.]+) ([-\d.]+)\)", location_str)
            if match:
                lon = float(match.group(1))
                lat = float(match.group(2))
                return lat, lon
            else:
                print(f"Failed to extract lat/long from location: {location_str}")
                return None, None
        except Exception as e:
            print(f"Error extracting lat/long: {e}")
            return None, None

    def convert_zulu_to_timezone(self, zulu_time_str, tz_offset, tz_name):
        if zulu_time_str is None:
            return ''
        try:
            if 'T' in zulu_time_str:
                if '.' in zulu_time_str:
                    zulu_time = datetime.strptime(zulu_time_str, '%Y-%m-%dT%H:%M:%S.%fZ')
                else:
                    zulu_time = datetime.strptime(zulu_time_str, '%Y-%m-%dT%H:%M:%SZ')
                local_time = zulu_time.replace(tzinfo=timezone.utc).astimezone(timezone(timedelta(hours=tz_offset)))
                return local_time.strftime('%Y-%m-%d %H:%M:%S')
        except ValueError as e:
            print(f"Failed to convert Zulu time to {tz_name}: {e}")
            return zulu_time_str

    def parse_reports(self, templates, reports, tz_offset, tz_name):
        def safe_value(value):
            if value is None or (isinstance(value, str) and value.strip() == ''):
                return ' '
            else:
                return value

        try:
            if self.output_option.get() == "Combined Workbook":
                workbook = Workbook()
                workbook.remove(workbook.active)

                for report_type, fields in templates.items():
                    sheet_name = ''.join([c for c in report_type if c.isalnum() or c in (' ', '_')]).strip()[:31]
                    if not sheet_name:
                        sheet_name = 'Report'
                    main_sheet = workbook.create_sheet(title=sheet_name)
                    main_sheet.append([field['csv_header'] for field in fields])

                    duplicate_sheet_name = f"{sheet_name}_duplicates"
                    duplicate_sheet = workbook.create_sheet(title=duplicate_sheet_name[:31])
                    duplicate_sheet.append([field['csv_header'] for field in fields])

                    self.process_reports_for_type(reports, report_type, fields, main_sheet, duplicate_sheet, tz_offset, tz_name)

                self.delete_columns(workbook)
                output_path = os.path.join(self.output_parent_folder, f"Exported TAK Reports {datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx")
                workbook.save(output_path)
                messagebox.showinfo("Success", f"Reports parsed and saved to {output_path}")

            else:
                for report_type, fields in templates.items():
                    workbook = Workbook()
                    workbook.remove(workbook.active)
                    sheet_name = ''.join([c for c in report_type if c.isalnum() or c in (' ', '_')]).strip()[:31]
                    if not sheet_name:
                        sheet_name = 'Report'

                    main_sheet = workbook.create_sheet(title=sheet_name)
                    main_sheet.append([field['csv_header'] for field in fields])

                    duplicate_sheet_name = f"{sheet_name}_duplicates"
                    duplicate_sheet = workbook.create_sheet(title=duplicate_sheet_name[:31])
                    duplicate_sheet.append([field['csv_header'] for field in fields])

                    self.process_reports_for_type(reports, report_type, fields, main_sheet, duplicate_sheet, tz_offset, tz_name)

                    self.delete_columns(workbook)
                    sanitized_report_type = ''.join([c for c in report_type if c.isalnum() or c in (' ', '_')]).strip()
                    output_filename = f"{sanitized_report_type} {datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
                    output_path = os.path.join(self.output_parent_folder, output_filename)
                    workbook.save(output_path)

                messagebox.showinfo("Success", f"Reports parsed and saved to separate workbooks in {self.output_parent_folder}")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while parsing reports: {e}")

    def return_to_home(self):
        self.destroy()
        Home_Page.open_home_page()

if __name__ == "__main__":
    app = TAKReportGUI()
